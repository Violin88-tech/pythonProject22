import os.path
from os.path import basename
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook


# Создаем  директорию resourses и проверяем что создался фаил.
def test_create_zip():
    if not os.path.exists("resourses/"):
        os.mkdir("resourses/")
    with ZipFile('resourses/zip.zip', "w") as zip_file:
        for folderName, subfolder, filenames in os.walk("tmp/"):
            for filename in filenames:
                file_path = os.path.join(folderName, filename)
                zip_file.write(file_path, basename(file_path))
    file_list = os.listdir("resourses/")
    assert "zip.zip" in file_list


# Просматриваем фаийлы в архиве и в папке tmp. Проверяем что название и их колличество равно.
def test_file_in_zip():
    with ZipFile('resourses/zip.zip', "r") as zip_file:
        file_list_in_zip = zip_file.namelist()
    file_list_in_folder = os.listdir("tmp/")
    assert file_list_in_zip == file_list_in_folder


# Читаем и проверяем что в csv фаиле есть Privet
def test_read_csv_file():
    with ZipFile('resourses/zip.zip') as zip_file:
        text = zip_file.read('csv.csv')
    assert text == b'Privet'

# Читаем и проверяем что в pdf фаиле есть PDF текст
def test_read_pdf_file():
    with ZipFile('resourses/zip.zip') as zip_file:
        with zip_file.open('file_pdf.pdf') as pdf:
            reader = PdfReader(pdf)
            assert "pytest Documentation" in reader.pages[0].extract_text()

# Читаем и проверяем что в xlsx фаиле есть текст Privet
def test_read_xlsx_file():
    with ZipFile("resourses/zip.zip") as zip_file:
        with zip_file.open('xlsx.xlsx') as xlsx:
            workbook = load_workbook("tmp/xlsx.xlsx")
            sheet = workbook.active
            a = sheet.cell(row=1, column=1).value
            assert a == 'Goodbye'